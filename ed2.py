import pandas as pd
from openpyxl import load_workbook


def firstTableRead(path):
    reader = pd.ExcelFile(path)
    dfR = reader.parse(0, header=None)
    osmi_red = dfR.iloc[7]
    dfR.columns = osmi_red
    dfR = dfR.iloc[8:].reset_index(drop=True)
    dfR.columns = dfR.columns.str.strip()
    return dfR

def secTableRead(path,sheetName):
    xls = pd.ExcelFile(path)
    df = xls.parse(sheetName)
    df.columns = df.columns.str.strip()
    return df

def update(df, dfR):
    merged_df = pd.merge(df, dfR, left_on="Fabr.broj aktivnog brojila", right_on="SERIJSKI BROJ BROJILA", how="left")

    df["Krajnje stanja aktiv. VT"] = merged_df["1.8.1 [kWh]"]
    df["Krajnje stanja aktiv. NT"] = merged_df["1.8.2 [kWh]"]
    df["Krajnje stanja reaktiv. VT"] = merged_df["3.8.1 [kVArh]"]
    df["Krajnje stanja reaktiv. NT"] = merged_df["3.8.2 [kVArh]"]
    df["Stanja maksigrafa"] = merged_df["1.6.1 [kW]"]


    df["Korekcija JT"] = merged_df["2.8.1 [kWh]"]
    df["Korekcija VT"] = merged_df["2.8.2 [kWh]"]
    df["Korekcija NT"] = merged_df["4.8.1 [kVArh]"]
    df["Korekcija VT_R"] = merged_df["4.8.2 [kVArh]"]
    df["Korekcija NT_R"] = merged_df["2.6.1 [kW]"]

    return df

def writeTable(df, original_path, output_path, sheet_name):
    wb = load_workbook(original_path)
    ws = wb[sheet_name]

    start_row = 2
    start_col = 1

    for i, row in df.iterrows():
        for j, value in enumerate(row):
            ws.cell(row=start_row + i, column=start_col + j, value=value)

    wb.save(output_path)




# Glavni deo
path1 = r"c:\Users\SS\Desktop\eksel za eps\radni\Nites AMM očitavanja - 01.01.2025.xlsx"
path2 = r"c:\Users\SS\Desktop\eksel za eps\radni\KP - Registar početna stanja i obračun - XII -  03.01.2025.xlsx"
output_path = r"2UPDATOVANA-KP - Registar početna stanja i obračun - XII -  03.01.2025.xlsx"
sheetName="Obracun KP za XII - GS"

dfR = firstTableRead(path1)
df = secTableRead(path2,sheetName)


dfUpdated = update(df, dfR)
writeTable(dfUpdated, path2, output_path,sheetName)